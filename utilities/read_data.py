import csv
import openpyxl

def getlogindata():
    test_LogInPage_data = [{"userID": "seattletester", "password": "Seattle123"}]
    return test_LogInPage_data

def getCSVData(fileName):
    # create an empty list to store rows
    rows = []
    # open the CSV file
    dataFile = open(fileName, "r")
    # create a CSV Reader from CSV file
    reader = csv.reader(dataFile)
    # skip the headers
    next(reader)
    # add rows from reader to list
    for row in reader:
        rows.append(row)
    return rows

def getExcelData(test_case_name):
    Dict = {}
    book = openpyxl.load_workbook("../testdata/testData.xlsx")
    sheet = book.active
    for i in range(1, sheet.max_row + 1): # to get rows
        if sheet.cell(row=i, column=1).value == test_case_name:
            for j in range(2, sheet.max_column + 1): # to get columns
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    return [Dict]